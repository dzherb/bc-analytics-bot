from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from models.participants import (
    Participant,
    ParticipantsReport,
    ParticipantType,
)
from services.export import (
    export_csv,
    export_excel,
)


def test_export_csv_with_export_date_header(tmp_path: Path) -> None:
    participants = [
        Participant(
            user_id='user123',
            username='testuser',
            full_name='Test User',
            about='About',
            registered_at=None,
            seen_as={ParticipantType.AUTHOR},
        )
    ]

    report = ParticipantsReport(
        exported_at=datetime(2024, 1, 1, 12, 0, 0),
        participants=participants,
    )

    file_path = tmp_path / 'participants.csv'
    export_csv(report, str(file_path), sep=',')

    content = file_path.read_text(encoding='utf-8').splitlines()

    assert content[0] == 'Дата экспорта,2024-01-01'
    assert content[1].startswith('Username,')
    assert 'Дата экспорта' not in content[1]


def test_export_excel_with_export_date_header(tmp_path: Path) -> None:
    participants = [
        Participant(
            user_id='user123',
            username='testuser',
            full_name='Test User',
            about='About',
            registered_at=None,
            seen_as={ParticipantType.AUTHOR},
        )
    ]

    report = ParticipantsReport(
        exported_at=datetime(2024, 1, 1, 12, 0, 0),
        participants=participants,
    )

    file_path = tmp_path / 'participants.xlsx'
    export_excel(report, str(file_path))

    workbook = load_workbook(file_path)
    worksheet = workbook.active

    assert worksheet is not None

    assert worksheet.cell(row=1, column=1).value == 'Дата экспорта'
    assert worksheet.cell(row=1, column=2).value == '2024-01-01'

    assert worksheet.cell(row=2, column=1).value == 'Username'
    assert worksheet.cell(row=2, column=5).value != 'Дата экспорта'
